#!/usr/bin/env python3
"""
CSV to Excel Converter

This script collects all CSV files in a directory, combines them into a single Excel file
with each CSV on a separate sheet, and formats URIs as clickable hyperlinks.

Usage:
    python csv_to_excel.py [directory_path]

Example:
    python csv_to_excel.py /path/to/data
"""

import os
import glob
import pandas as pd
import argparse
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.colors import Color


def collect_csv_files(directory_path='.'):
    """
    Find all CSV files in the specified directory.
    
    Args:
        directory_path: Path to the directory to search (default: current directory)
        
    Returns:
        A list of paths to CSV files
    """
    # Ensure directory path ends with separator
    if not directory_path.endswith(os.sep) and directory_path != '.':
        directory_path += os.sep
    
    # Find all CSV files in the directory
    pattern = os.path.join(directory_path, '*.csv')
    csv_files = glob.glob(pattern)
    
    return csv_files


def create_excel_with_hyperlinks(csv_files, output_file='collected_terms.xlsx'):
    """
    Create an Excel file with each CSV on a separate sheet and format URIs as hyperlinks.
    
    Args:
        csv_files: List of CSV file paths
        output_file: Name of the output Excel file
    """
    if not csv_files:
        print("No CSV files found.")
        return
    
    # Create an Excel writer object
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        for csv_file in csv_files:
            # Read the CSV file
            try:
                df = pd.read_csv(csv_file)
                
                # Check if 'URI' column exists
                if 'URI' not in df.columns:
                    print(f"Warning: No 'URI' column found in {csv_file}. Skipping hyperlink formatting.")
                
                # Get the base name of the file (without extension) to use as sheet name
                sheet_name = os.path.splitext(os.path.basename(csv_file))[0]
                
                # Excel has a 31 character limit for sheet names
                if len(sheet_name) > 31:
                    sheet_name = sheet_name[:31]
                
                # Write the dataframe to the Excel file
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                print(f"Added {csv_file} to sheet '{sheet_name}'")
            except Exception as e:
                print(f"Error processing {csv_file}: {e}")
    
    # Now add hyperlinks using openpyxl
    wb = load_workbook(output_file)
    
    for csv_file in csv_files:
        sheet_name = os.path.splitext(os.path.basename(csv_file))[0]
        if len(sheet_name) > 31:
            sheet_name = sheet_name[:31]
        
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Find the URI column
            uri_col = None
            for col_idx, cell in enumerate(ws[1]):
                if cell.value == 'URI':
                    uri_col = col_idx + 1  # openpyxl uses 1-based indexing
                    break
            
            if uri_col:
                # Format URIs as hyperlinks
                for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    cell = row[uri_col - 1]
                    if cell.value:
                        # Set the hyperlink
                        cell.hyperlink = cell.value
                        # Format the cell as a hyperlink (blue, underlined)
                        cell.font = Font(color="0000FF", underline="single")
    
    # Save the workbook
    wb.save(output_file)
    print(f"Excel file saved as {output_file} with hyperlinks.")


def main():
    """Main function to handle command line arguments and run the script."""
    parser = argparse.ArgumentParser(description="Convert CSV files to Excel with hyperlinks")
    parser.add_argument("directory", nargs='?', default='.',
                        help="Directory to search for CSV files (default: current directory)")
    parser.add_argument("--output", "-o", default='collected_terms.xlsx',
                        help="Output Excel file name (default: collected_terms.xlsx)")
    
    args = parser.parse_args()
    
    try:
        csv_files = collect_csv_files(args.directory)
        
        if csv_files:
            print(f"Found {len(csv_files)} CSV files:")
            for file in csv_files:
                print(f"  - {file}")
            
            create_excel_with_hyperlinks(csv_files, args.output)
        else:
            print(f"No CSV files found in {args.directory}")
            
    except Exception as e:
        print(f"Error: {e}")


if __name__ == "__main__":
    main()
